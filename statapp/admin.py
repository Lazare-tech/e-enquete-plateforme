from django.contrib import admin
from .models import Category, StatFile, UserFileAccess,FAQ,StatVariable,VariableCategory,Document,ContactMessage
from django.contrib.auth.models import User
from django import forms
from adminsortable2.admin import SortableAdminMixin
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
import csv
from django.utils.html import format_html  
from openpyxl import Workbook
import openpyxl
##############################################################################
##############################################################################
def export_to_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="export_fichiers.csv"'
    writer = csv.writer(response)
    # Entêtes
    writer.writerow(['Titre', 'Catégorie', 'Auteur', 'Téléchargements', 'Date création'])
    # Données
    for obj in queryset:
        writer.writerow([obj.title, obj.category.title, obj.created_by.username if obj.created_by else "N/A", obj.download_count, obj.created_at])
    return response

export_to_csv.short_description = "Exporter en CSV"
###
def export_statfiles_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fichiers Statistiques"

    # Style pour l'entête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal="center")

    # Définition des colonnes
    headers = [
        'Titre de l\'enquête', 
        'Catégorie', 
        'Auteur', 
        'Format', 
        'Téléchargements', 
        'Date d\'ajout', 
        'Lien du fichier'
    ]
    ws.append(headers)

    # Appliquer le style aux entêtes
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Remplissage des données
    for obj in queryset.select_related('category', 'created_by'):
        # On récupère l'extension via la méthode que tu as dans ton modèle
        ext = obj.get_extension() if hasattr(obj, 'get_extension') else "N/A"
        
        ws.append([
            obj.title,
            obj.category.title,
            obj.created_by.username if obj.created_by else "Système",
            ext.upper(),
            obj.download_count,
            obj.created_at.strftime("%d/%m/%Y %H:%M"),
            request.build_absolute_uri(obj.file.url) if obj.file else "Pas de fichier"
        ])

    # Ajustement automatique de la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # Préparation de la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="export_fichiers_statistiques.xlsx"'
    wb.save(response)
    return response

export_statfiles_excel.short_description = "Exporter la sélection en format Excel"
###

def export_variables_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Variables Statistiques"

    # Style des entêtes
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    bold_font = Font(bold=True)

    # Entêtes
    headers = ['Catégorie', 'Label (Nom)', 'Valeur', 'Ajouté par', 'Date']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font

    # Données (triées par catégorie pour plus de clarté)
    for var in queryset.select_related('category', 'created_by').order_by('category__title'):
        ws.append([
            var.category.title,
            var.label,
            var.value,
            var.created_by.username if var.created_by else "N/A",
            var.created_at.strftime("%d/%m/%Y")
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="export_variables.xlsx"'
    wb.save(response)
    return response

export_variables_excel.short_description = "Exporter les variables en Excel"

##############################################################################
##############################################################################
# 1. Création d'un formulaire personnalisé pour l'accès
class UserFileAccessForm(forms.ModelForm):
    # On redéfinit le champ user pour changer son étiquette
    user = forms.ModelChoiceField(
        queryset=User.objects.all(),
        label="Utilisateur (Email)"
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # On remplace l'affichage par défaut (username) par l'email
        self.fields['user'].label_from_instance = lambda obj: f"{obj.email} ({obj.username})" if obj.email else obj.username

    class Meta:
        model = UserFileAccess
        fields = '__all__'
# 1. Permet d'ajouter/voir les utilisateurs autorisés directement dans la page du fichier
class UserFileAccessInline(admin.TabularInline):
    model = UserFileAccess
    extra = 1  # Affiche une ligne vide par défaut pour ajouter un utilisateur rapidement

@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ('title', 'slug', 'icon')
    # list_editable = ('is_active',) 
    prepopulated_fields = {'slug': ('title',)}
    search_fields = ('title',)

@admin.register(StatFile)
class StatFileAdmin(admin.ModelAdmin):
    actions = [export_to_csv, export_statfiles_excel]
    # Ajout du slug dans l'affichage
    list_display = ('title', 'category', 'created_by', 'slug', 'download_count', 'is_active', 'created_at')
    list_filter = ('category', 'is_active', 'created_at')
    search_fields = ('title', 'description')
    list_editable = ('is_active',)
    
    # Génère le slug automatiquement pendant que tu tapes le titre
    prepopulated_fields = {'slug': ('title',)}
    
    # Intégration de l'Inline pour gérer les accès WhatsApp en un clic
    inlines = [UserFileAccessInline]

    fieldsets = (
        ('Informations Générales', {
            'fields': ('category', 'title', 'slug', 'description','created_by', 'is_active', 'is_protected')
        }),
        ('Fichier & Données', {
            'fields': ('file', 'download_count'),
        }),
        ('Dates (Lecture seule)', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',), # Cache cette section par défaut
        }),
    )
    
    readonly_fields = ('download_count', 'created_at', 'updated_at')

# 2. Enregistrer aussi le modèle d'accès séparément pour une vue d'ensemble
@admin.register(UserFileAccess)
class UserFileAccessAdmin(admin.ModelAdmin):
    form = UserFileAccessForm # On applique le formulaire ici
    list_display = ('user_email', 'stat_file', 'granted_at')
    list_filter = ('granted_at', 'stat_file')
    search_fields = ('user__email', 'user__username', 'stat_file__title')
    #
    def user_email(self, obj):
        return obj.user.email
    user_email.short_description = 'Email de l\'utilisateur'
#####
@admin.register(FAQ)
class FAQAdmin(SortableAdminMixin, admin.ModelAdmin):
    list_display = ('order', 'question', 'is_active')
    list_editable = ('is_active',) # Permet de masquer une question d'un clic
    search_fields = ('question', 'answer')
#####
@admin.register(VariableCategory)
class VariableCategoryAdmin(admin.ModelAdmin):
    list_display = ('title', 'is_active', 'created_at', 'variable_count')
    list_editable = ('is_active',) # Validez d'un clic dans la liste !
    list_filter = ('is_active',)
    search_fields = ('title',)

    def variable_count(self, obj):
        return obj.variables.count()
    variable_count.short_description = "Nombre de variables"

@admin.register(StatVariable)
class StatVariableAdmin(admin.ModelAdmin):
    actions = [export_variables_excel]
    list_display = ('label', 'value', 'category', 'created_by')
    list_filter = ('category', 'category__is_active')
    search_fields = ('label', 'value')
    # Permet de changer de catégorie directement dans la liste
    list_editable = ('category',)
#####################################################################
@admin.register(Document)
class DocumentAdmin(admin.ModelAdmin):
    # Colonnes visibles dans la liste des documents
    list_display = ('titre', 'ajoute_par', 'date_upload', 'apercu_fichier')
    list_filter = ('date_upload', 'ajoute_par')
    search_fields = ('titre', 'description')
    
    # On cache les champs automatiques pour ne pas encombrer le formulaire
    exclude = ('ajoute_par', 'slug')

    def save_model(self, request, obj, form, change):
        # On enregistre l'utilisateur qui a fait l'upload
        if not obj.pk:
            obj.ajoute_par = request.user
        super().save_model(request, obj, form, change)

    def apercu_fichier(self, obj):
        if obj.fichier:
            return format_html(
                '<a href="{}" target="_blank" style="font-weight:bold; color:#122046;">'
                '<i class="fas fa-external-link-alt"></i> Ouvrir le document'
                '</a>',
                obj.fichier.url
            )
        return "Aucun fichier"
    
    apercu_fichier.allow_tags = True # Autorise le HTML pour le lien
    apercu_fichier.short_description = "Lien de téléchargement"
########
@admin.register(ContactMessage)
class ContactMessageAdmin(admin.ModelAdmin):
    # Colonnes affichées dans la liste
    list_display = ('nom', 'email', 'objet', 'date_envoi')
    
    # Filtres sur le côté droit
    list_filter = ('date_envoi', 'objet')
    
    # Barre de recherche (très utile quand tu auras beaucoup de messages)
    search_fields = ('nom', 'email', 'objet', 'message')
    
    # Tri par défaut (les plus récents en premier)
    ordering = ('-date_envoi',)
    
    # Rendre les champs en lecture seule pour éviter les modifs accidentelles
    readonly_fields = ('nom', 'email', 'telephone', 'objet', 'message', 'date_envoi')

    # Optionnel : Empêcher l'ajout manuel depuis l'admin (les messages viennent du site)
    def has_add_permission(self, request):
        return False