from django.contrib import admin
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from datetime import datetime
# Register your models here.
from django.contrib import admin
from .models import Survey, Question, Option, Submission, Answer
from adminsortable2.admin import SortableAdminMixin, SortableTabularInline
from django.utils.html import format_html
from django.urls import reverse
# 1. Permet d'ajouter des options directement sous la question
class OptionInline(SortableTabularInline):
    model = Option
    extra = 1

# 2. Permet d'ajouter des questions directement sous le questionnaire
class QuestionInline(admin.StackedInline):
    model = Question
    extra = 0
    show_change_link = True
    # On organise les champs pour que la logique de saut soit bien visible
    fieldsets = (
        ('Configuration de base', {
            'fields': (('question_type', 'order'), 'label', 'help_text', 'required')
        }),
        ('Logique de saut (Optionnel)', {
            'fields': (('depends_on', 'dependency_value'),),
            'description': "Affiche cette question uniquement si une réponse spécifique est donnée à une autre.",
            'classes': ('collapse',), # Cache cette section par défaut pour ne pas encombrer
        }),
    )
    
    # Pour pouvoir sélectionner facilement la question parente
    autocomplete_fields = ['depends_on']
    class Media:
        js = ('js/admin_enq.js',) # Chemin vers ton fichier JS
###########################################
def export_survey_to_csv(modeladmin, request, queryset):
    # On prend la première enquête sélectionnée (ou on peut boucler pour faire un ZIP)
    survey = queryset.first()
    if not survey:
        return
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"export_{survey.id}_{datetime.now().strftime('%Y%m%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Correction pour les caractères accentués en Excel
    response.write(u'\ufeff'.encode('utf8'))
    writer = csv.writer(response)
    
    # En-têtes
    questions = survey.questions.exclude(question_type__in=['Entete', 'section']).order_by('order')
    headers = ['ID Soumission', 'Date', 'Enquêteur'] + [q.label for q in questions]
    writer.writerow(headers)
    
    # URL de base pour les fichiers média
    base_url = request.build_absolute_uri('/')[:-1]
    
    submissions = Submission.objects.filter(survey=survey).prefetch_related('answers')
    
    for sub in submissions:
        row = [
            sub.id, 
            sub.submitted_at.strftime("%d/%m/%Y %H:%M"), 
            sub.enumerator.username if sub.enumerator else "Anonyme"
        ]
        
        # Dictionnaire des réponses (Texte ou URL complète pour fichier)
        answers_dict = {}
        for a in sub.answers.all():
            if a.file_upload:
                answers_dict[a.question_id] = f"{base_url}{a.file_upload.url}"
            else:
                answers_dict[a.question_id] = a.value
        
        for q in questions:
            row.append(answers_dict.get(q.id, ""))
        writer.writerow(row)
        
    return response

export_survey_to_csv.short_description = "Exporter en CSV"
def export_survey_to_excel(modeladmin, request, queryset):
    survey = queryset.first()
    if not survey:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Réponses"
    
    # Style Shine Agency pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="122046", end_color="122046", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    questions = survey.questions.exclude(question_type__in=['Entete', 'section']).order_by('order')
    headers = ['ID', 'Date', 'Enquêteur'] + [q.label for q in questions]
    ws.append(headers)
    
    # Application des styles
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    base_url = request.build_absolute_uri('/')[:-1]
    submissions = Submission.objects.filter(survey=survey).prefetch_related('answers')
    
    for sub in submissions:
        row = [
            sub.id, 
            sub.submitted_at.strftime("%d/%m/%Y %H:%M"), 
            sub.enumerator.username if sub.enumerator else "Anonyme"
        ]
        
        answers_dict = {}
        for a in sub.answers.all():
            if a.file_upload:
                answers_dict[a.question_id] = f"{base_url}{a.file_upload.url}"
            else:
                answers_dict[a.question_id] = a.value
                
        for q in questions:
            row.append(answers_dict.get(q.id, ""))
        ws.append(row)

    # Ajustement automatique de la largeur des colonnes
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"export_{survey.id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

export_survey_to_excel.short_description = "Exporter en Excel (.xlsx)"
@admin.register(Survey)
class SurveyAdmin(admin.ModelAdmin):
    list_display = ('title','share_link', 'created_by', 'is_active', 'created_at')
    list_filter = ('is_active', 'created_at')
    search_fields = ('title',)
    actions = [export_survey_to_csv, export_survey_to_excel] # Ajout de l'action ici
    inlines = [QuestionInline] # On injecte les questions dans le formulaire Survey
    #
    def share_link(self, obj):
        # On construit l'URL complète
        url = reverse('survey_builder:fill', kwargs={'uid': obj.uid})
        # Note : En admin, on utilise souvent le domaine par défaut ou request
        full_url = f"http://127.0.0.1:8000{url}" # À adapter selon ton domaine de prod
        
        # Le HTML du bouton "Copier" avec un petit script intégré
        return format_html(
            '''
            <div style="display:flex; align-items:center;">
                <input type="text" value="{}" id="url_{}" style="width:1px; opacity:0; position:absolute;">
                <button type="button" onclick="copyToClipboard('url_{}')" 
                        style="background:#122046; color:white; border:none; padding:4px 8px; border-radius:4px; cursor:pointer; font-size:10px;">
                    <i class="fas fa-copy"></i> Copier le lien
                </button>
                <script>
                    function copyToClipboard(id) {{
                        var copyText = document.getElementById(id);
                        navigator.clipboard.writeText(copyText.value);
                        alert("Lien copié dans le presse-papier !");
                    }}
                </script>
            </div>
            ''',
            full_url, obj.id, obj.id
        )

    share_link.short_description = "Partage"
    class Media:
        js = ('js/admin_cp.js',)
    ####
 
from django.contrib.admin import SimpleListFilter

@admin.register(Question)
class QuestionAdmin(SortableAdminMixin, admin.ModelAdmin):
    # 1. On affiche les colonnes importantes
    list_display = ('label', 'survey_name', 'order', 'question_type')
    
    # 2. RÉACTIVATION DE LA RECHERCHE
    # Permet de chercher par le texte de la question OU le titre du formulaire
    search_fields = ('label', 'survey__title')
    
    # 3. FILTRES LATERAUX
    list_filter = (
        ('survey', admin.RelatedOnlyFieldListFilter),
        'question_type',
    )
    
    # 4. GESTION DES CHOIX (Radio/Checkbox)
    # C'est ici que tu peux ajouter tes options (Choix unique/multiple)
    inlines = [OptionInline]
    
    ordering = ('survey', 'order')

    def survey_name(self, obj):
        return obj.survey.title
    survey_name.short_description = "Formulaire"
# Visualisation des réponses (Lecture seule pour la sécurité des données)
# 1. Définir l'Inline à l'extérieur
class AnswerInline(admin.TabularInline):
    model = Answer
    fields = ('question', 'display_value')
    readonly_fields = ('question', 'display_value')
    extra = 0
    can_delete = False # Sécurité : on ne supprime pas les réponses ici

    def display_value(self, obj):
        # Vérification si un fichier est présent
        if obj.file_upload:
            url = obj.file_upload.url
            if obj.question.question_type == 'image':
                # Miniature cliquable pour les images
                return format_html(
                    '<a href="{0}" target="_blank">'
                    '<img src="{0}" style="width: 80px; height: auto; border-radius: 5px; border: 1px solid #ddd;"/>'
                    '</a>', url
                )
            elif obj.question.question_type == 'audio':
                # Lien avec icône pour l'audio
                return format_html(
                    '<a href="{}" target="_blank" style="font-weight:bold; color:#122046; text-decoration:none;">'
                    '<i class="fas fa-play-circle"></i> 🎵 Écouter l\'audio'
                    '</a>', url
                )
        
        # Affichage du texte classique (text, select, checkbox)
        return obj.value or "-"

    display_value.short_description = "Réponse / Fichier"

# 2. Enregistrer l'Admin de Soumission
@admin.register(Submission)
class SubmissionAdmin(admin.ModelAdmin):
    list_display = ('survey', 'enumerator', 'submitted_at')
    list_filter = ('survey', 'submitted_at', 'enumerator') # Filtres utiles
    readonly_fields = ('survey', 'enumerator', 'submitted_at')
    
    inlines = [AnswerInline]